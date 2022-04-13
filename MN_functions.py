import re, sys, json
import time, random#, testlink
import openpyxl
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import NoSuchElementException, WebDriverException
from selenium.webdriver.support import expected_conditions as EC
from random import choice
import pathlib
from pathlib import Path
import os
from sys import platform
import MN_function

chrome_options = webdriver.ChromeOptions()


#result=open(os.path.dirname(Path(__file__).absolute())+'\\result.txt','w')

class bcolors:
    HEADER = '\033[95m'
    OKBLUE = '\033[94m'
    OKGREEN = '\033[92m'
    WARNING = '\033[93m'
    FAIL = '\033[91m'
    ENDC = '\033[0m'
    BOLD = '\033[1m'
    UNDERLINE = '\033[4m'

#testcase_fail = bcolors.FAIL + ">>> Test case status: fail" + bcolors.ENDC
#testcase_pass = bcolors.OKGREEN + ">>> Test case status: pass" + bcolors.ENDC

#print(json_file)
class objects:
    now = datetime.now()
    year = now.strftime("%Y")
    month = now.strftime("%m")
    day = now.strftime("%d")
    time1 = now.strftime("%H:%M:%S")
    date_time = now.strftime("%Y/%m/%d, %H:%M:%S")
    date_id = date_time.replace("/", "").replace(", ", "").replace(":", "")[2:]
    testcase_pass = "Test case status: pass"
    testcase_fail = "Test case status: fail"

if platform == "linux" or platform == "linux2":
    local_path = "/home/oem/groupware-auto-test"
    json_file = local_path + "/MN_groupware_auto.json"
    with open(json_file) as json_data_file:
        data = json.load(json_data_file)
    driver = webdriver.Chrome("/usr/bin/chromedriver")
    log_folder = "/Log/"
    log_testcase = "/testcase_log/"
    execution_log = local_path + log_folder + "execution_log_" + str(objects.date_id) + ".txt"
    fail_log = execution_log.replace("execution_log_", "fail_log_")
    error_log = execution_log.replace("execution_log_", "error_log_")
    testcase_log = local_path + log_testcase + "MN_testcase_result_" + str(objects.date_id) + ".xlsx"

    #import common_functions
    #driver = common_functions.driver
    
else:
    #local_path = "D:\\Ngoc\\ngoc_automationtest"
    local_path = "C:\\Users\\Ngoc\\Desktop\\ngoc_automationtest"
    json_file = local_path + "\\MN_groupware_auto.json"
    with open(json_file) as json_file:
        data = json.load(json_file)
    driver = webdriver.Chrome(local_path + "\\chromedriver.exe")
    log_folder = "\\Log\\"
    log_testcase = "\\testcase_log\\"
    execution_log = local_path + log_folder + "execution_log_" + str(objects.date_id) + ".txt"
    fail_log = execution_log.replace("execution_log_", "fail_log_")
    error_log = execution_log.replace("execution_log_", "error_log_")
    testcase_log = local_path + log_testcase + "MN_testcase_result_" + str(objects.date_id) + ".xlsx"
    
driver.maximize_window()
driver.implicitly_wait(10)

'''
# create log file of fail test case
open(execution_log, "x").close()

# create log file of fail test case
open(fail_log, "x").close()

# create log file of fail test case
open(error_log, "x").close()
'''

logs = [execution_log, fail_log, error_log, testcase_log]
for log in logs:
    if ".txt" in log:
        open(log, "x").close()
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(row=1, column=1).value = "Menu"
        ws.cell(row=1, column=2).value = "Sub-Menu"
        ws.cell(row=1, column=3).value = "Test Case Name"
        ws.cell(row=1, column=4).value = "Status"
        ws.cell(row=1, column=5).value = "Description"
        ws.cell(row=1, column=6).value = "Date"
        ws.cell(row=1, column=7).value = "Tester"
        wb.save(log)

def Logging(msg):
    print(msg)
    log_msg = open(execution_log, "a")
    log_msg.write(str(msg) + "\n")
    log_msg.close()

def ValidateFailResultAndSystem(fail_msg):
    print(fail_msg)
    append_fail_result = open(fail_log, "a")
    append_fail_result.write("[FAILED TEST CASE] " + str(fail_msg) + "\n")
    append_fail_result.close()

def TestCase_LogResult(menu, sub_menu, testcase, status, description, tester):
    #Logging(description)

    if status == "Pass":
        print(objects.testcase_pass)
    else:
        print(objects.testcase_fail)
    
    wb = openpyxl.load_workbook(testcase_log)
    current_sheet = wb.active
    start_row = len(list(current_sheet.rows)) + 1

    current_sheet.cell(row=start_row, column=1).value = menu
    current_sheet.cell(row=start_row, column=2).value = sub_menu
    current_sheet.cell(row=start_row, column=3).value = testcase
    current_sheet.cell(row=start_row, column=4).value = status
    current_sheet.cell(row=start_row, column=5).value = description
    current_sheet.cell(row=start_row, column=6).value = objects.date_time
    current_sheet.cell(row=start_row, column=7).value = tester

    wb.save(testcase_log)

'''TESTLINK_API_PYTHON_SERVER_URL= 'http://qa1.hanbiro.net/testlink/lib/api/xmlrpc/v1/xmlrpc.php'
TESTLINK_API_PYTHON_DEVKEY= 'e80ce28cbff80bd624cc2679c915596d'

tls = testlink.TestLinkHelper(TESTLINK_API_PYTHON_SERVER_URL, TESTLINK_API_PYTHON_DEVKEY).connect(testlink.TestlinkAPIClient)

def TestlinkResult_Pass(external_id):
    tls.reportTCResult(testcaseexternalid=external_id, testplanid=7377, buildname="v3.8.46", status='p', notes='Test Case [' + external_id + '] passed')

def TestlinkResult_Fail(external_id):
    tls.reportTCResult(testcaseexternalid=external_id, testplanid=7377, buildname="v3.8.46", status='f', notes='Test Case [' + external_id + '] failed')'''

def access_qa(domain):
    driver.get(domain)
    print(domain)
    print("- Access login page")
    driver.find_element_by_id("log-userid").send_keys(data["user_name"])
    print("- Input user ID")
    frame_element = driver.find_element_by_id("iframeLoginPassword")
    driver.switch_to.frame(frame_element)
    driver.find_element_by_id("p").send_keys(data["user_password"])
    driver.switch_to.default_content()
    print("- Input user password")
    driver.find_element_by_id("btn-log").send_keys(Keys.RETURN)
    print("- Click button Login")

def close_pop_up():
    try:
        WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH, data["turn_off"]))).send_keys(Keys.RETURN)
        print("- Close pop up Successfully")
    except:
        print("- Don't show pop up")
